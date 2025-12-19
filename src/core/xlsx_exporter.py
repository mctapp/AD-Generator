# core/xlsx_exporter.py
# 스프레드시트 내보내기

import os
from typing import List

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
    from openpyxl.utils import get_column_letter
    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False

from .pdf_parser import ScriptEntry


class XLSXExporter:
    """XLSX 스프레드시트 내보내기"""
    
    def __init__(self):
        if not HAS_OPENPYXL:
            raise ImportError("openpyxl 패키지가 필요합니다: pip install openpyxl")
    
    def export(self, entries: List[ScriptEntry], output_path: str, 
               include_brackets: bool = True):
        """
        ScriptEntry 리스트를 XLSX로 내보내기
        
        Args:
            entries: ScriptEntry 리스트
            output_path: 저장 경로
            include_brackets: 괄호 내용 컬럼 포함 여부
        """
        wb = Workbook()
        ws = wb.active
        ws.title = "음성해설 대본"
        
        # 스타일 정의
        header_font = Font(bold=True, size=11, color="FFFFFF")
        header_fill = PatternFill(start_color="4F46E5", end_color="4F46E5", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center")
        
        cell_alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
        tc_alignment = Alignment(horizontal="center", vertical="center")
        
        thin_border = Border(
            left=Side(style='thin', color='E5E5E5'),
            right=Side(style='thin', color='E5E5E5'),
            top=Side(style='thin', color='E5E5E5'),
            bottom=Side(style='thin', color='E5E5E5')
        )
        
        # 헤더
        if include_brackets:
            headers = ["#", "타임코드", "원본 TC", "지시사항", "음성해설 대본"]
            col_widths = [6, 15, 10, 20, 60]
        else:
            headers = ["#", "타임코드", "원본 TC", "음성해설 대본"]
            col_widths = [6, 15, 10, 70]
        
        for col, (header, width) in enumerate(zip(headers, col_widths), 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = thin_border
            ws.column_dimensions[get_column_letter(col)].width = width
        
        # 데이터 행
        for row, entry in enumerate(entries, 2):
            if include_brackets:
                data = [
                    entry.index,
                    entry.timecode_formatted,
                    entry.timecode_raw,
                    entry.bracket_content,
                    entry.script_text
                ]
            else:
                data = [
                    entry.index,
                    entry.timecode_formatted,
                    entry.timecode_raw,
                    entry.script_text
                ]
            
            for col, value in enumerate(data, 1):
                cell = ws.cell(row=row, column=col, value=value)
                cell.border = thin_border
                
                if col <= 3:  # 번호, 타임코드
                    cell.alignment = tc_alignment
                else:
                    cell.alignment = cell_alignment
            
            # 행 높이 조정
            ws.row_dimensions[row].height = max(20, len(entry.script_text) // 40 * 15 + 20)
        
        # 첫 행 고정
        ws.freeze_panes = 'A2'
        
        wb.save(output_path)
    
    def export_sync_report(self, sync_entries: List, output_path: str):
        """
        동기화 결과를 XLSX로 내보내기
        """
        wb = Workbook()
        ws = wb.active
        ws.title = "SRT 동기화 결과"
        
        # 스타일
        header_font = Font(bold=True, size=11, color="FFFFFF")
        header_fill = PatternFill(start_color="10B981", end_color="10B981", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center")
        
        thin_border = Border(
            left=Side(style='thin', color='E5E5E5'),
            right=Side(style='thin', color='E5E5E5'),
            top=Side(style='thin', color='E5E5E5'),
            bottom=Side(style='thin', color='E5E5E5')
        )
        
        # 상태별 색상
        status_fills = {
            'synced': PatternFill(start_color="D1FAE5", end_color="D1FAE5", fill_type="solid"),
            'shorter': PatternFill(start_color="FEF3C7", end_color="FEF3C7", fill_type="solid"),
            'longer': PatternFill(start_color="FECACA", end_color="FECACA", fill_type="solid"),
            'missing': PatternFill(start_color="F3F4F6", end_color="F3F4F6", fill_type="solid"),
        }
        
        # 헤더
        headers = ["#", "타임코드", "원본 길이", "WAV 길이", "차이", "상태", "파일명"]
        col_widths = [6, 15, 12, 12, 12, 10, 25]
        
        for col, (header, width) in enumerate(zip(headers, col_widths), 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = thin_border
            ws.column_dimensions[get_column_letter(col)].width = width
        
        # 데이터 행
        for row, entry in enumerate(sync_entries, 2):
            original_dur = entry.original_end_ms - entry.start_ms
            
            data = [
                entry.index,
                self._ms_to_display(entry.start_ms),
                f"{original_dur/1000:.1f}초",
                f"{entry.wav_duration_ms/1000:.1f}초" if entry.wav_duration_ms else "-",
                f"{entry.diff_ms/1000:+.1f}초" if entry.status != 'missing' else "-",
                self._status_to_korean(entry.status),
                entry.wav_filename
            ]
            
            for col, value in enumerate(data, 1):
                cell = ws.cell(row=row, column=col, value=value)
                cell.border = thin_border
                cell.alignment = Alignment(horizontal="center", vertical="center")
                
                # 상태에 따른 행 색상
                if col == 6:  # 상태 컬럼
                    cell.fill = status_fills.get(entry.status, status_fills['synced'])
        
        # 첫 행 고정
        ws.freeze_panes = 'A2'
        
        wb.save(output_path)
    
    def _ms_to_display(self, ms: int) -> str:
        """밀리초를 표시용 시간으로 변환"""
        hours = ms // 3600000
        minutes = (ms % 3600000) // 60000
        seconds = (ms % 60000) // 1000
        return f"{hours:02d}:{minutes:02d}:{seconds:02d}"
    
    def _status_to_korean(self, status: str) -> str:
        """상태를 한글로 변환"""
        return {
            'synced': '✓ 일치',
            'shorter': '▼ 짧음',
            'longer': '▲ 김',
            'missing': '- 없음'
        }.get(status, status)
