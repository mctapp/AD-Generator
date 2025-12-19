# core/srt_sync.py
# SRT-WAV 동기화 모듈

import os
import re
import wave
from dataclasses import dataclass
from typing import List, Optional, Tuple


@dataclass
class SyncEntry:
    """동기화 항목"""
    index: int
    start_ms: int
    original_end_ms: int
    wav_duration_ms: int
    synced_end_ms: int
    text: str
    wav_filename: str
    status: str  # 'synced', 'shorter', 'longer', 'missing'
    diff_ms: int


class SRTSync:
    """SRT-WAV 동기화"""
    
    def __init__(self, fps: float = 24):
        self.fps = fps
        self.entries: List[SyncEntry] = []
    
    def parse_srt(self, srt_path: str) -> List[dict]:
        """SRT 파일 파싱"""
        entries = []
        
        with open(srt_path, 'r', encoding='utf-8') as f:
            content = f.read()
        
        # SRT 블록 패턴
        pattern = re.compile(
            r'(\d+)\n'
            r'(\d{2}:\d{2}:\d{2}[,\.]\d{3})\s*-->\s*(\d{2}:\d{2}:\d{2}[,\.]\d{3})\n'
            r'(.*?)(?=\n\n|\n*$)',
            re.DOTALL
        )
        
        for match in pattern.finditer(content):
            index = int(match.group(1))
            start_time = match.group(2)
            end_time = match.group(3)
            text = match.group(4).strip()
            
            entries.append({
                'index': index,
                'start_ms': self._srt_time_to_ms(start_time),
                'end_ms': self._srt_time_to_ms(end_time),
                'text': text
            })
        
        return entries
    
    def _srt_time_to_ms(self, time_str: str) -> int:
        """SRT 시간 형식을 밀리초로 변환"""
        time_str = time_str.replace(',', '.')
        parts = time_str.split(':')
        hours = int(parts[0])
        minutes = int(parts[1])
        sec_parts = parts[2].split('.')
        seconds = int(sec_parts[0])
        ms = int(sec_parts[1]) if len(sec_parts) > 1 else 0
        
        return hours * 3600000 + minutes * 60000 + seconds * 1000 + ms
    
    def _ms_to_srt_time(self, ms: int) -> str:
        """밀리초를 SRT 시간 형식으로 변환"""
        hours = ms // 3600000
        minutes = (ms % 3600000) // 60000
        seconds = (ms % 60000) // 1000
        milliseconds = ms % 1000
        return f"{hours:02d}:{minutes:02d}:{seconds:02d},{milliseconds:03d}"
    
    def _ms_to_filename_tc(self, ms: int) -> str:
        """밀리초를 파일명용 타임코드로 변환"""
        hours = ms // 3600000
        minutes = (ms % 3600000) // 60000
        seconds = (ms % 60000) // 1000
        frames = int((ms % 1000) / 1000 * self.fps)
        return f"{hours:02d}_{minutes:02d}_{seconds:02d}_{frames:02d}"
    
    def get_wav_duration(self, wav_path: str) -> Optional[int]:
        """WAV 파일 길이(밀리초) 반환"""
        try:
            with wave.open(wav_path, 'rb') as wav_file:
                frames = wav_file.getnframes()
                rate = wav_file.getframerate()
                duration_ms = int(frames / rate * 1000)
                return duration_ms
        except Exception:
            return None
    
    def analyze(self, srt_path: str, wav_folder: str) -> List[SyncEntry]:
        """
        SRT와 WAV 파일 분석
        
        Args:
            srt_path: SRT 파일 경로
            wav_folder: WAV 파일 폴더 경로
        
        Returns:
            SyncEntry 리스트
        """
        self.entries = []
        srt_entries = self.parse_srt(srt_path)
        
        for entry in srt_entries:
            # WAV 파일명 생성 (타임코드 기반)
            tc_filename = self._ms_to_filename_tc(entry['start_ms'])
            wav_filename = f"{tc_filename}.wav"
            wav_path = os.path.join(wav_folder, wav_filename)
            
            # WAV 길이 확인
            wav_duration = self.get_wav_duration(wav_path) if os.path.exists(wav_path) else None
            
            if wav_duration is None:
                sync_entry = SyncEntry(
                    index=entry['index'],
                    start_ms=entry['start_ms'],
                    original_end_ms=entry['end_ms'],
                    wav_duration_ms=0,
                    synced_end_ms=entry['end_ms'],
                    text=entry['text'],
                    wav_filename=wav_filename,
                    status='missing',
                    diff_ms=0
                )
            else:
                original_duration = entry['end_ms'] - entry['start_ms']
                diff = wav_duration - original_duration
                synced_end = entry['start_ms'] + wav_duration
                
                if abs(diff) < 100:  # 100ms 미만 차이는 동기화됨
                    status = 'synced'
                elif diff < 0:
                    status = 'shorter'
                else:
                    status = 'longer'
                
                sync_entry = SyncEntry(
                    index=entry['index'],
                    start_ms=entry['start_ms'],
                    original_end_ms=entry['end_ms'],
                    wav_duration_ms=wav_duration,
                    synced_end_ms=synced_end,
                    text=entry['text'],
                    wav_filename=wav_filename,
                    status=status,
                    diff_ms=diff
                )
            
            self.entries.append(sync_entry)
        
        return self.entries
    
    def generate_synced_srt(self) -> str:
        """동기화된 SRT 생성"""
        srt_content = []
        
        for entry in self.entries:
            if entry.status == 'missing':
                end_ms = entry.original_end_ms
            else:
                end_ms = entry.synced_end_ms
            
            srt_block = f"{entry.index}\n"
            srt_block += f"{self._ms_to_srt_time(entry.start_ms)} --> {self._ms_to_srt_time(end_ms)}\n"
            srt_block += f"{entry.text}\n"
            
            srt_content.append(srt_block)
        
        return '\n'.join(srt_content)
    
    def save_synced_srt(self, output_path: str):
        """동기화된 SRT 저장"""
        content = self.generate_synced_srt()
        with open(output_path, 'w', encoding='utf-8') as f:
            f.write(content)
    
    def get_summary(self) -> dict:
        """분석 결과 요약"""
        total = len(self.entries)
        synced = sum(1 for e in self.entries if e.status == 'synced')
        shorter = sum(1 for e in self.entries if e.status == 'shorter')
        longer = sum(1 for e in self.entries if e.status == 'longer')
        missing = sum(1 for e in self.entries if e.status == 'missing')
        
        return {
            'total': total,
            'synced': synced,
            'shorter': shorter,
            'longer': longer,
            'missing': missing
        }
    
    def set_fps(self, fps: float):
        """FPS 설정"""
        self.fps = fps
    
    @property
    def results(self) -> List[SyncEntry]:
        """entries의 별칭 (호환성)"""
        return self.entries
    
    def save_report_xlsx(self, output_path: str):
        """XLSX 리포트 저장"""
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Font, Alignment, PatternFill
            
            wb = Workbook()
            ws = wb.active
            ws.title = "동기화 리포트"
            
            # 헤더
            headers = ["#", "타임코드", "원본길이(ms)", "WAV길이(ms)", "차이(ms)", "상태"]
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col, value=header)
                cell.font = Font(bold=True)
                cell.alignment = Alignment(horizontal='center')
            
            # 데이터
            for row, entry in enumerate(self.entries, 2):
                ws.cell(row=row, column=1, value=entry.index)
                ws.cell(row=row, column=2, value=self._ms_to_filename_tc(entry.start_ms).replace('_', ':'))
                ws.cell(row=row, column=3, value=entry.original_end_ms - entry.start_ms)
                ws.cell(row=row, column=4, value=entry.wav_duration_ms)
                ws.cell(row=row, column=5, value=entry.diff_ms)
                
                status_map = {'synced': 'OK', 'shorter': '여유', 'longer': '초과', 'missing': '누락'}
                ws.cell(row=row, column=6, value=status_map.get(entry.status, entry.status))
            
            wb.save(output_path)
            return True
        except Exception:
            return False
    
    def save_report_txt(self, output_path: str):
        """TXT 리포트 저장"""
        lines = ["동기화 리포트", "=" * 50, ""]
        
        for entry in self.entries:
            tc = self._ms_to_filename_tc(entry.start_ms).replace('_', ':')
            status_map = {'synced': 'OK', 'shorter': '여유', 'longer': '초과', 'missing': '누락'}
            status = status_map.get(entry.status, entry.status)
            
            lines.append(f"#{entry.index} [{tc}] {status}")
            lines.append(f"  원본: {entry.original_end_ms - entry.start_ms}ms, WAV: {entry.wav_duration_ms}ms, 차이: {entry.diff_ms}ms")
            lines.append("")
        
        summary = self.get_summary()
        lines.append("=" * 50)
        lines.append(f"총계: {summary['total']}개")
        lines.append(f"  OK: {summary['synced']}, 여유: {summary['shorter']}, 초과: {summary['longer']}, 누락: {summary['missing']}")
        
        with open(output_path, 'w', encoding='utf-8') as f:
            f.write('\n'.join(lines))
        return True
